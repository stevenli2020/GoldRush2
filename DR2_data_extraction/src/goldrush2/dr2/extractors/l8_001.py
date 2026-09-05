"""DR2 extractor for L8-001: monthly global gold ETF net flows."""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from goldrush2.dr2.collectors import wgc
from goldrush2.dr2.extractors._wgc_common import HORIZON_LOOKBACKS, build_output, degraded, finite, parse_date

VARIABLE_ID = "L8-001"
SOURCE_NAME = "WGC ETF Flows - Global Gold ETF Net Flows (tonnes)"
SOURCE_URL = "https://www.gold.org/"
from goldrush2.paths import DR2_ROOT as PROJECT_ROOT
RAW_DIR = PROJECT_ROOT / "data" / "raw" / "wgc"
OUTPUT_PATH = PROJECT_ROOT / "data" / "current" / f"{VARIABLE_ID}.json"


def parse_flows_workbook(path: Path) -> list[dict[str, Any]]:
    """Parse the WGC Demand by month sheet by summing per-fund flow columns."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Demand by month" not in workbook.sheetnames:
        raise ValueError("WGC Demand by month sheet was not found")
    rows = list(workbook["Demand by month"].iter_rows(values_only=True))
    header_index = next((index for index, row in enumerate(rows) if any(str(value).strip().lower() == "date" for value in row if value is not None) and any(str(value).strip().lower() == "value (usd)" for value in row if value is not None)), None)
    if header_index is None:
        raise ValueError("WGC flow header was not found")
    headers = [str(value).strip().lower() if value is not None else "" for value in rows[header_index]]
    date_index = next(index for index, value in enumerate(headers) if value == "date")
    fund_start = next(index for index, value in enumerate(headers) if value == "value (usd)") + 1
    observations: list[dict[str, Any]] = []
    seen_dates: set[str] = set()
    for row in rows[header_index + 1 :]:
        observation_date = parse_date(row[date_index] if date_index < len(row) else None)
        if observation_date is None or observation_date in seen_dates:
            continue
        values = [float(value) for value in row[fund_start:] if finite(value)]
        if not values:
            continue
        seen_dates.add(observation_date)
        observations.append({"date": observation_date, "value": round(sum(values), 10)})
    if not observations:
        raise ValueError("WGC flow sheet contained no valid observations")
    return observations


def build_degraded_output(summary: str, *, as_of_date: str | None = None) -> dict[str, Any]:
    """Build a zero-confidence output for a WGC collection failure."""
    return {"variable_id": VARIABLE_ID, "as_of_date": as_of_date or date.today().isoformat(), "source_name": SOURCE_NAME, "source_url": SOURCE_URL, "data_frequency": "Monthly", "observation_date": None, "horizons": {horizon: degraded(summary) for horizon in HORIZON_LOOKBACKS}}


def run(*, output_path: Path = OUTPUT_PATH, raw_dir: Path = RAW_DIR) -> dict[str, Any]:
    """Collect, parse, and write the current L8-001 output."""
    workbook = wgc.fetch_wgc_workbook(raw_dir)
    if workbook is None:
        summary = "STALE DATA — WGC ETF workbook is unavailable and no fresh cache exists." if wgc.LAST_FETCH_STALE else "SOURCE UNAVAILABLE — WGC ETF workbook could not be downloaded."
        output = build_degraded_output(summary)
    else:
        try:
            output = build_output(VARIABLE_ID, SOURCE_NAME, SOURCE_URL, parse_flows_workbook(workbook), cached=wgc.LAST_FETCH_USED_CACHE, value_label="Gold ETF net flow")
        except (OSError, ValueError, KeyError) as exc:
            output = build_degraded_output(f"EXTRACTION FAILED — {exc}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(output, indent=2) + "\n", encoding="utf-8")
    return output


def main() -> None:
    print(json.dumps(run(), indent=2))


if __name__ == "__main__":
    main()
