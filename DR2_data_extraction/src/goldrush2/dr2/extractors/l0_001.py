"""DR2 extractor for WGC annual above-ground gold stocks."""

from __future__ import annotations

import json
import math
from datetime import date
from pathlib import Path
from typing import Any

from goldrush2.dr2.collectors.wgc import LAST_FETCH_STALE, LAST_FETCH_USED_CACHE, fetch_wgc_above_ground_stocks
from goldrush2.paths import DR2_ROOT as PROJECT_ROOT

VARIABLE_ID = "L0-001"
SOURCE_NAME = "WGC GoldHub - Above-Ground Gold Stock (tonnes)"
SOURCE_URL = "https://www.gold.org/goldhub/data/how-much-gold"
PARSED_CACHE_PATH = PROJECT_ROOT / "data/cache/wgc/l0_001.json"
OUTPUT_PATH = PROJECT_ROOT / "data/current/L0-001.json"


def _finite(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def parse_above_ground_workbook(path: Path, *, cache_path: Path | None = PARSED_CACHE_PATH) -> list[dict[str, Any]]:
    """Extract the ``Total`` row from WGC's annual stocks sheet."""
    path = Path(path)
    stat = path.stat()
    if cache_path is not None:
        try:
            cached = json.loads(Path(cache_path).read_text(encoding="utf-8"))
            if cached.get("source_file") == path.name and cached.get("source_mtime_ns") == stat.st_mtime_ns and cached.get("source_size") == stat.st_size and isinstance(cached.get("observations"), list):
                return cached["observations"]
        except (OSError, ValueError, json.JSONDecodeError):
            pass
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = "Above-ground stocks"
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"required sheet missing: {sheet_name}")
    rows = list(workbook[sheet_name].iter_rows(values_only=True))
    if not rows or not str(rows[0][0]).lower().startswith("above-ground stocks"):
        raise ValueError("above-ground stocks title is missing")
    # Locate the row containing annual year labels instead of assuming a
    # fixed spreadsheet offset; WGC may insert title or methodology rows.
    header_index = next((index for index, row in enumerate(rows) if sum(1 for value in row if value is not None and str(value).strip().isdigit() and 1900 <= int(float(value)) <= 2100) >= 2), None)
    if header_index is None:
        raise ValueError("annual year header was not found")
    header = rows[header_index]
    total_row = next((row for row in rows[header_index + 1 :] if row and str(row[0]).strip().lower() == "total"), None)
    if total_row is None:
        raise ValueError("Total above-ground stock row was not found")
    observations: dict[str, dict[str, Any]] = {}
    for index, year in enumerate(header):
        if year is None or str(year).strip() == "":
            continue
        try:
            year_number = int(float(year))
        except (TypeError, ValueError):
            # WGC may append a non-annual YTD column (for example YTD'26*).
            # L0-001 is annual, so ignore that column rather than rejecting
            # the otherwise valid historical series.
            continue
        value = _finite(total_row[index] if index < len(total_row) else None)
        if value is None or value < 0:
            raise ValueError(f"above-ground stock must be non-negative for {year_number}")
        observed = f"{year_number:04d}-12-31"
        if observed in observations:
            raise ValueError(f"duplicate annual observation: {observed}")
        observations[observed] = {"date": observed, "value": value}
    result = [observations[key] for key in sorted(observations)]
    if not result:
        raise ValueError("workbook contains no annual above-ground stock observations")
    if cache_path is not None:
        target = Path(cache_path)
        target.parent.mkdir(parents=True, exist_ok=True)
        payload = {"source_file": path.name, "source_mtime_ns": stat.st_mtime_ns, "source_size": stat.st_size, "observations": result}
        temporary = target.with_suffix(target.suffix + ".tmp")
        temporary.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
        temporary.replace(target)
    return result


def build_output(observations: list[dict[str, Any]], *, cached: bool = False) -> dict[str, Any]:
    rows = sorted(observations, key=lambda row: str(row["date"]))
    current = rows[-1] if rows else None
    horizons: dict[str, Any] = {
        "1-5d": {"signal": 0, "confidence": 1, "evidence": {"summary": "Annual data does not support 1-5d horizon."}},
        "1-3m": {"signal": 0, "confidence": 1, "evidence": {"summary": "Annual data does not support 1-3m horizon."}},
    }
    for horizon, lookback in (("1-3y", 3), ("3-10y", 7)):
        target_year = int(str(current["date"])[:4]) - lookback if current else None
        comparison = next((row for row in rows if current and row["date"] == f"{target_year:04d}-12-31"), None)
        if current is None or comparison is None:
            summary = f"MISSING DATA — no observation exactly {lookback} years prior."
            if cached or LAST_FETCH_STALE:
                summary += " SOURCE UNAVAILABLE — cached data used."
            data = {"current_stock": current["value"] if current else None, "current_date": current["date"] if current else None, "comparison_stock": None, "comparison_date": None, "change_tonnes": None, "change_pct": None, f"{lookback}_years_ago_year": target_year}
            horizons[horizon] = {"signal": None, "confidence": 0, "evidence": {"data": data, "error": f"No data exactly {lookback} years prior", "current_year": int(str(current["date"])[:4]) if current else None, "summary": summary}}
            continue
        change = round(float(current["value"]) - float(comparison["value"]), 10)
        change_pct = round(change / float(comparison["value"]) * 100, 10) if comparison["value"] else None
        signal = 1 if change > 0 else -1 if change < 0 else 0
        direction = "rose" if signal == 1 else "fell" if signal == -1 else "was unchanged"
        summary = f"Above-ground gold stock {direction} by {abs(change):,.0f} tonnes ({change_pct:+.2f}%) compared to {lookback} years ago, {'bullish' if signal == 1 else 'bearish' if signal == -1 else 'neutral'} for gold."
        if cached or LAST_FETCH_USED_CACHE:
            summary += " SOURCE UNAVAILABLE — cached data used."
        data = {"current_stock": float(current["value"]), "current_date": current["date"], "comparison_stock": float(comparison["value"]), "comparison_date": comparison["date"], "change_tonnes": change, "change_pct": change_pct, f"{lookback}_years_ago_year": target_year}
        horizons[horizon] = {"signal": signal, "confidence": 1, "evidence": {"data": data, "summary": summary}}
    return {"variable_id": VARIABLE_ID, "data_frequency": "Annual", "source_name": SOURCE_NAME, "source_url": SOURCE_URL, "observation_date": current["date"] if current else None, "as_of_date": date.today().isoformat(), "horizons": horizons}


def run(*, workbook: Path | None = None, output_path: Path = OUTPUT_PATH, raw_dir: Path = PROJECT_ROOT / "data/raw/wgc") -> dict[str, Any]:
    source = Path(workbook) if workbook else fetch_wgc_above_ground_stocks(raw_dir)
    cached = LAST_FETCH_USED_CACHE
    if source is None:
        fallback = Path(raw_dir) / "above-ground-gold-stocks.xlsx"
        if fallback.exists():
            source, cached = fallback, True
    if source is None:
        raise RuntimeError("WGC above-ground stocks workbook is unavailable")
    output = build_output(parse_above_ground_workbook(source), cached=cached)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = output_path.with_suffix(output_path.suffix + ".tmp")
    temporary.write_text(json.dumps(output, indent=2) + "\n", encoding="utf-8")
    temporary.replace(output_path)
    return output


if __name__ == "__main__":
    run()
