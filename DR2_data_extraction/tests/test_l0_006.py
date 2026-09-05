"""Tests for the quarterly WGC GDT recycling extractor."""

import pytest
from openpyxl import Workbook

from goldrush2.dr2.extractors import l0_006


def observations(count=21, current=50.0, comparison=40.0):
    rows = [{"date": f"{2020 + i // 4:04d}-{(i % 4 + 1) * 3:02d}-30", "value": 45.0} for i in range(count)]
    if count > 4:
        rows[-1 - 4]["value"] = comparison
    rows[-1]["value"] = current
    return rows


@pytest.mark.parametrize("horizon, lookback", [("1-3y", 4), ("3-10y", 20)])
@pytest.mark.parametrize("current, comparison, signal", [(50.0, 40.0, -1), (30.0, 40.0, 1), (40.0, 40.0, 0)])
def test_signal_directions_and_lookbacks(horizon, lookback, current, comparison, signal):
    rows = observations(current=current)
    rows[-1 - lookback]["value"] = comparison
    result = l0_006.build_output(rows)
    assert result["horizons"][horizon]["signal"] == signal
    assert result["horizons"][horizon]["confidence"] == 1


def test_short_horizons_are_inapplicable():
    result = l0_006.build_output(observations())
    for horizon in ("1-5d", "1-3m"):
        assert result["horizons"][horizon]["signal"] == 0
        assert result["horizons"][horizon]["confidence"] == 1


def gdt_workbook(tmp_path, *, negative=False, supply_sheet=True):
    workbook = Workbook()
    balance = workbook.active
    balance.title = "Gold Balance"
    for _ in range(4):
        balance.append([])
    balance.append([None, None, None, "Q1'25", "Q2'25"])
    balance.append([None, "Total Bar and Coin", None, 100.0, 100.0])
    balance.append([None, "Bars", None, 60.0, 60.0])
    balance.append([None, "Official Coins", None, 30.0, 30.0])
    balance.append([None, "Medals Imitation Coins", None, 10.0, 10.0])
    if not supply_sheet:
        balance.append([None, "Recycled Gold", None, -1.0 if negative else 25.0, 26.0])
    else:
        supply = workbook.create_sheet("Supply")
        for _ in range(4):
            supply.append([])
        supply.append([None, None, None, "Q1'25", "Q2'25"])
        supply.append([None, "Recycled Gold", None, -1.0 if negative else 25.0, 26.0])
    path = tmp_path / "GDT_Tables_Q1'25_EN.xlsx"
    workbook.save(path)
    return path


def test_dynamic_sheet_row_and_quarter_normalization(tmp_path):
    result = l0_006.parse_recycling_workbook(gdt_workbook(tmp_path), cache_path=tmp_path / "parsed.json")
    assert result == [{"date": "2025-03-31", "value": 25.0}, {"date": "2025-06-30", "value": 26.0}]


def test_negative_global_recycling_rejected(tmp_path):
    with pytest.raises(ValueError, match="Negative"):
        l0_006.parse_recycling_workbook(gdt_workbook(tmp_path, negative=True), cache_path=None)


def test_recycling_cache_reuse(tmp_path, monkeypatch):
    workbook = gdt_workbook(tmp_path)
    cache = tmp_path / "parsed.json"
    expected = l0_006.parse_recycling_workbook(workbook, cache_path=cache)
    monkeypatch.setattr("openpyxl.load_workbook", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("reparsed")))
    assert l0_006.parse_recycling_workbook(workbook, cache_path=cache) == expected


def test_insufficient_history():
    assert l0_006.build_output(observations(4))["horizons"]["1-3y"]["confidence"] == 0


def test_schema():
    result = l0_006.build_output(observations())
    assert result["variable_id"] == "L0-006"
    assert result["data_frequency"] == "Quarterly"


def test_source_failure(monkeypatch, tmp_path):
    monkeypatch.setattr(l0_006.wgc, "fetch_wgc_gdt_workbook", lambda path: None)
    monkeypatch.setattr(l0_006.wgc, "LAST_FETCH_STALE", False)
    result = l0_006.run(output_path=tmp_path / "out.json", raw_dir=tmp_path)
    assert result["horizons"]["1-3y"]["confidence"] == 0
    assert "SOURCE UNAVAILABLE" in result["horizons"]["1-3y"]["evidence"]["summary"]


def test_stale_cache(monkeypatch, tmp_path):
    monkeypatch.setattr(l0_006.wgc, "fetch_wgc_gdt_workbook", lambda path: None)
    monkeypatch.setattr(l0_006.wgc, "LAST_FETCH_STALE", True)
    result = l0_006.run(output_path=tmp_path / "out.json", raw_dir=tmp_path)
    assert "STALE DATA" in result["horizons"]["1-3y"]["evidence"]["summary"]
