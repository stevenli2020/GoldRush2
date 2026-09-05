from datetime import date

import pytest
from openpyxl import Workbook, load_workbook

from goldrush2.dr2.extractors.l0_001 import build_output, parse_above_ground_workbook


def make_book(path, values=None, *, sheet="Above-ground stocks"):
    book = Workbook()
    ws = book.active
    ws.title = sheet
    ws.append(["Above-ground stocks (tonnes)"])
    ws.append([])
    years = list(range(2010, 2020))
    ws.append([None, *years])
    ws.append(["Jewellery", *([1] * len(years))])
    ws.append(["Total", *(values or list(range(100, 110)))])
    book.save(path)


def observations(count=10):
    return [{"date": f"{2010 + i}-12-31", "value": float(100 + i)} for i in range(count)]


def test_sheet_parsing(tmp_path):
    path = tmp_path / "stocks.xlsx"
    make_book(path, [100, 101] + [102] * 8)
    result = parse_above_ground_workbook(path, cache_path=None)
    assert result[0] == {"date": "2010-12-31", "value": 100.0}


def test_cache_reuse(tmp_path):
    path, cache = tmp_path / "stocks.xlsx", tmp_path / "l0_001.json"
    make_book(path)
    first = parse_above_ground_workbook(path, cache_path=cache)
    assert parse_above_ground_workbook(path, cache_path=cache) == first


def test_missing_sheet(tmp_path):
    path = tmp_path / "bad.xlsx"
    make_book(path, sheet="Other")
    with pytest.raises(ValueError, match="required sheet"):
        parse_above_ground_workbook(path, cache_path=None)


def test_missing_total(tmp_path):
    path = tmp_path / "bad.xlsx"
    make_book(path)
    book = load_workbook(path)
    book.active["A5"] = "Other"
    book.save(path)
    with pytest.raises(ValueError, match="Total"):
        parse_above_ground_workbook(path, cache_path=None)


@pytest.mark.parametrize("value", [-1, "abc", None, float("nan"), float("inf")])
def test_invalid_stock_rejected(tmp_path, value):
    path = tmp_path / "bad.xlsx"
    make_book(path, [value] + [100] * 9)
    with pytest.raises(ValueError):
        parse_above_ground_workbook(path, cache_path=None)


@pytest.mark.parametrize("current,comparison,signal", [(120, 100, 1), (80, 100, -1), (100, 100, 0)])
def test_signal_direction(current, comparison, signal):
    rows = observations(8)
    rows[-1]["value"] = current
    rows[-4]["value"] = comparison
    assert build_output(rows)["horizons"]["1-3y"]["signal"] == signal


@pytest.mark.parametrize("horizon", ["1-5d", "1-3m"])
def test_short_horizons(horizon):
    output = build_output(observations())
    assert output["horizons"][horizon]["signal"] == 0
    assert output["horizons"][horizon]["confidence"] == 1


def test_three_and_seven_year_lookbacks():
    output = build_output(observations(10))
    assert output["horizons"]["1-3y"]["evidence"]["data"]["comparison_date"] == "2016-12-31"
    assert output["horizons"]["3-10y"]["evidence"]["data"]["comparison_date"] == "2012-12-31"
    assert output["horizons"]["1-3y"]["evidence"]["data"]["3_years_ago_year"] == 2016


def test_gap_does_not_use_nearest_observation():
    rows = [row for row in observations(10) if row["date"] != "2016-12-31"]
    output = build_output(rows)
    horizon = output["horizons"]["1-3y"]
    assert horizon["signal"] is None and horizon["confidence"] == 0
    assert horizon["evidence"]["error"] == "No data exactly 3 years prior"
    assert horizon["evidence"]["current_year"] == 2019


def test_insufficient_history():
    output = build_output(observations(3))
    assert output["horizons"]["1-3y"]["confidence"] == 0


def test_cached_degradation():
    output = build_output(observations(3), cached=True)
    assert "SOURCE UNAVAILABLE" in output["horizons"]["1-3y"]["evidence"]["summary"]


def test_change_percentage_and_schema():
    output = build_output(observations(8))
    data = output["horizons"]["1-3y"]["evidence"]["data"]
    assert data["change_tonnes"] == 3 and data["change_pct"] == pytest.approx(3 / 104 * 100)
    assert output["variable_id"] == "L0-001" and set(output["horizons"]) == {"1-5d", "1-3m", "1-3y", "3-10y"}
