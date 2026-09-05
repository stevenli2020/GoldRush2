from datetime import date, timedelta
import json

import pytest
from openpyxl import Workbook

from goldrush2.dr2.extractors.l9_001 import LOOKBACKS, build_output, parse_premiums_workbook


def make_book(path, values=None, *, title="Chinese Premium/Discount (US$/oz) - 5 day moving average"):
    book = Workbook()
    sheet = book.active
    sheet.title = "Chinese premiums-discounts"
    sheet.append(["Gold market premium/discount data"])
    sheet.append(["methodology"])
    sheet.append([])
    sheet.append([title])
    sheet.append(["Source: WGC"])
    for index, value in enumerate(values or [1, 2, 3], start=0):
        sheet.append([(date(2020, 1, 3) + timedelta(days=index * 7)).isoformat(), value])
    book.save(path)


def observations(count=10):
    return [{"date": (date(2020, 1, 3) + timedelta(days=i * 7)).isoformat(), "value": float(i), "premium_usd": float(i), "premium_pct": None} for i in range(count)]


def test_exact_sheet_and_rows(tmp_path):
    path = tmp_path / "gold-premiums.xlsx"
    make_book(path, [1.5, -2.0])
    rows = parse_premiums_workbook(path, cache_path=None)
    assert rows[0]["date"] == "2020-01-03" and rows[1]["premium_usd"] == -2


def test_parsed_cache_reused(tmp_path):
    path, cache = tmp_path / "gold-premiums.xlsx", tmp_path / "gold_premiums.json"
    make_book(path, [1])
    first = parse_premiums_workbook(path, cache_path=cache)
    second = parse_premiums_workbook(path, cache_path=cache)
    assert first == second and json.loads(cache.read_text())["observations"] == first


def test_missing_sheet_rejected(tmp_path):
    path = tmp_path / "x.xlsx"
    Workbook().save(path)
    with pytest.raises(ValueError, match="required sheet"):
        parse_premiums_workbook(path, cache_path=None)


def test_bad_title_rejected(tmp_path):
    path = tmp_path / "x.xlsx"
    make_book(path, [1], title="wrong")
    with pytest.raises(ValueError, match="title"):
        parse_premiums_workbook(path, cache_path=None)


@pytest.mark.parametrize("value", [None, "abc", float("nan"), float("inf")])
def test_invalid_values_rejected(tmp_path, value):
    path = tmp_path / "x.xlsx"
    make_book(path, [value])
    with pytest.raises(ValueError):
        parse_premiums_workbook(path, cache_path=None)


def test_duplicate_dates_rejected(tmp_path):
    path = tmp_path / "x.xlsx"
    make_book(path, [1, 2])
    from openpyxl import load_workbook
    book = load_workbook(path)
    book.active["A7"] = book.active["A6"].value
    book.save(path)
    with pytest.raises(ValueError, match="duplicate"):
        parse_premiums_workbook(path, cache_path=None)


@pytest.mark.parametrize("current,comparison,signal", [(2, 1, 1), (1, 2, -1), (1, 1, 0)])
def test_signal_directions(current, comparison, signal):
    rows = observations(6)
    rows[-1]["premium_usd"] = current
    rows[-6]["premium_usd"] = comparison
    assert build_output(rows)["horizons"]["1-5d"]["signal"] == signal


@pytest.mark.parametrize("horizon,lookback", LOOKBACKS.items())
def test_lookbacks(horizon, lookback):
    rows = observations(lookback + 1)
    assert build_output(rows)["horizons"][horizon]["confidence"] == 1


def test_insufficient_history_is_zero_confidence():
    output = build_output(observations(5))
    assert output["horizons"]["1-5d"]["confidence"] == 0


def test_absolute_level_preserved():
    output = build_output(observations(6))
    data = output["horizons"]["1-5d"]["evidence"]["data"]
    assert data["current_value_usd"] == 5 and data["absolute_level"] == "positive"


def test_percentage_formula_when_columns_present(tmp_path):
    path = tmp_path / "x.xlsx"
    make_book(path, [1])
    from openpyxl import load_workbook
    book = load_workbook(path)
    book.active["C6"], book.active["D6"] = 110, 100
    book.save(path)
    row = parse_premiums_workbook(path, cache_path=None)[0]
    assert row["premium_pct"] == pytest.approx(10)


def test_output_schema():
    output = build_output(observations(6))
    assert output["variable_id"] == "L9-001"
    assert set(output["horizons"]) == set(LOOKBACKS)


def test_negative_absolute_level():
    rows = observations(6)
    rows[-1]["premium_usd"] = -1
    assert build_output(rows)["horizons"]["1-5d"]["evidence"]["data"]["absolute_level"] == "negative"
