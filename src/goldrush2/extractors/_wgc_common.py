"""Shared helpers for monthly World Gold Council ETF extractors."""

from __future__ import annotations

import json
import math
import re
import calendar
from datetime import date, datetime
from pathlib import Path
from typing import Any

HORIZON_LOOKBACKS = {"1-5d": 5, "1-3m": 63, "1-3y": 252, "3-10y": 756}
QUARTER_HORIZON_LOOKBACKS = {"1-3y": 4, "3-10y": 20}


def finite(value: Any) -> bool:
    """Return whether a value can be represented as a finite float."""
    try:
        return math.isfinite(float(value))
    except (TypeError, ValueError):
        return False


def parse_date(value: Any) -> str | None:
    """Normalize an Excel date or ISO-like string to YYYY-MM-DD."""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if value is None:
        return None
    text = str(value).strip()
    for parser in (date.fromisoformat,):
        try:
            return parser(text).isoformat()
        except ValueError:
            pass
    for pattern in ("%b %Y", "%B %Y"):
        try:
            return datetime.strptime(text, pattern).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def parse_official_changes_workbook(path: Any) -> list[dict[str, Any]]:
    """Read canonical country/date changes from the WGC Monthly sheet."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Monthly" not in workbook.sheetnames:
        raise ValueError("WGC Monthly sheet was not found")
    rows = list(workbook["Monthly"].iter_rows(values_only=True))
    header_index = next((index for index, row in enumerate(rows) if len(row) > 1 and str(row[1]).strip().lower() == "country"), None)
    if header_index is None:
        raise ValueError("WGC Monthly country header was not found")
    dates = [(index, parse_date(value)) for index, value in enumerate(rows[header_index]) if index >= 3 and parse_date(value) is not None]
    if not dates:
        raise ValueError("WGC Monthly date columns were not found")
    records: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        country = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if not country or country.lower() == "nan" or country.endswith("*"):
            continue
        for column, observation_date in dates:
            value = row[column] if column < len(row) else None
            if value in (None, "") or not finite(value):
                continue
            records.append({"country": country, "date": observation_date, "value": float(value)})
    if not records:
        raise ValueError("WGC Monthly sheet contained no canonical numeric changes")
    return records


def cumulative_net_change_series(path: Any, *, cache_path: Any | None = None) -> list[dict[str, Any]]:
    """Build a monthly cumulative index from canonical WGC country changes.

    The official-changes workbook reports signed monthly changes rather than
    a level.  This fallback starts at zero at the first reported month and
    cumulatively sums the global canonical-country changes thereafter.
    """
    source_path = Path(path)
    source_stat = source_path.stat()
    if cache_path is not None:
        parsed_cache = Path(cache_path)
        try:
            cached = json.loads(parsed_cache.read_text(encoding="utf-8"))
            if (cached.get("source_file") == source_path.name and cached.get("source_mtime_ns") == source_stat.st_mtime_ns and cached.get("source_size") == source_stat.st_size and isinstance(cached.get("observations"), list)):
                return cached["observations"]
        except (OSError, UnicodeDecodeError, json.JSONDecodeError, AttributeError):
            pass

    totals: dict[str, float] = {}
    for record in parse_official_changes_workbook(path):
        observation_date = str(record["date"])
        totals[observation_date] = totals.get(observation_date, 0.0) + float(record["value"])
    if not totals:
        raise ValueError("WGC Monthly sheet contained no dated changes")
    cumulative = 0.0
    series: list[dict[str, Any]] = []
    for observation_date in sorted(totals):
        cumulative += totals[observation_date]
        series.append({"date": observation_date, "value": round(cumulative, 10)})
    if cache_path is not None:
        parsed_cache = Path(cache_path)
        parsed_cache.parent.mkdir(parents=True, exist_ok=True)
        payload = {"source_file": source_path.name, "source_mtime_ns": source_stat.st_mtime_ns, "source_size": source_stat.st_size, "observations": series}
        temporary = parsed_cache.with_suffix(parsed_cache.suffix + ".tmp")
        temporary.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
        temporary.replace(parsed_cache)
    return series


def parse_quarter_label(value: Any) -> tuple[int, int] | None:
    """Parse common WGC quarter labels and return ``(year, quarter)``."""
    text = re.sub(r"\s+", " ", str(value).strip())
    match = re.fullmatch(r"Q([1-4])\s*['’]?\s*(\d{2}|\d{4})", text, re.IGNORECASE)
    if match:
        year = int(match.group(2))
        return (2000 + year if year < 100 else year, int(match.group(1)))
    match = re.fullmatch(r"(\d{4})\s*Q([1-4])", text, re.IGNORECASE)
    return (int(match.group(1)), int(match.group(2))) if match else None


def quarter_end_date(year: int, quarter: int) -> str:
    """Return the ISO quarter-end date for a year and quarter number."""
    month = quarter * 3
    return f"{year:04d}-{month:02d}-{calendar.monthrange(year, month)[1]:02d}"


def parse_gdt_quarterly_workbook(path: Any, *, cache_path: Any | None = None) -> dict[str, list[dict[str, Any]]]:
    """Parse quarterly bar-and-coin demand and recycled supply from a WGC GDT workbook."""
    source_path = Path(path)
    source_stat = source_path.stat()
    if cache_path is not None:
        parsed_cache = Path(cache_path)
        try:
            cached = json.loads(parsed_cache.read_text(encoding="utf-8"))
            if (cached.get("source_file") == source_path.name and cached.get("source_mtime_ns") == source_stat.st_mtime_ns and cached.get("source_size") == source_stat.st_size and isinstance(cached.get("observations"), dict)):
                return cached["observations"]
        except (OSError, UnicodeDecodeError, json.JSONDecodeError, AttributeError):
            pass

    from openpyxl import load_workbook

    workbook = load_workbook(source_path, read_only=True, data_only=True)

    def rows_for(sheet_name: str) -> list[tuple[Any, ...]]:
        if sheet_name not in workbook.sheetnames:
            return []
        return list(workbook[sheet_name].iter_rows(values_only=True))

    def quarter_columns(rows: list[tuple[Any, ...]]) -> dict[int, str]:
        for row in rows[:12]:
            columns: dict[int, str] = {}
            for index, value in enumerate(row):
                parsed = parse_quarter_label(value)
                if parsed:
                    columns[index] = quarter_end_date(*parsed)
            if columns:
                return columns
        return {}

    def find_row(rows: list[tuple[Any, ...]], predicate: Any) -> tuple[Any, ...] | None:
        for row in rows:
            labels = [str(value).strip() for value in row[:3] if isinstance(value, str) and value.strip()]
            if any(predicate(label) for label in labels):
                return row
        return None

    balance_rows = rows_for("Gold Balance")
    columns = quarter_columns(balance_rows)
    if not columns:
        raise ValueError("WGC GDT quarterly header was not found")
    total_row = find_row(balance_rows, lambda label: label.strip().lower() == "total bar and coin")
    bars_row = find_row(balance_rows, lambda label: label.strip().lower() == "bars")
    coins_row = find_row(balance_rows, lambda label: label.strip().lower() == "official coins")
    medals_row = find_row(balance_rows, lambda label: label.strip().lower() == "medals imitation coins")
    if not all((total_row, bars_row, coins_row, medals_row)):
        raise ValueError("WGC GDT bar-and-coin component rows were not found")
    demand: list[dict[str, Any]] = []
    for column, observation_date in columns.items():
        values = [row[column] if column < len(row) else None for row in (total_row, bars_row, coins_row, medals_row)]
        if not all(finite(value) for value in values):
            continue
        total, bars, coins, medals = (float(value) for value in values)
        if min(total, bars, coins, medals) < 0:
            raise ValueError(f"Negative WGC GDT bar-and-coin value for {observation_date}")
        if abs((bars + coins + medals) - total) > max(0.5, abs(total) * 0.01):
            raise ValueError(f"WGC GDT bar-and-coin components do not reconcile for {observation_date}")
        demand.append({"date": observation_date, "value": round(total, 10), "bars": round(bars, 10), "official_coins": round(coins, 10), "medals_imitation_coins": round(medals, 10)})
    if not demand:
        raise ValueError("WGC GDT workbook contained no quarterly bar-and-coin demand")

    recycling: list[dict[str, Any]] = []
    recycling_row = None
    recycling_columns = columns
    for sheet_name in ("Gold Balance", "Supply", "Supply_and_Demand"):
        rows = rows_for(sheet_name)
        candidate = find_row(rows, lambda label: "recycled gold" in label.lower())
        if candidate is not None:
            recycling_row, recycling_columns = candidate, quarter_columns(rows)
            break
    if recycling_row is None or not recycling_columns:
        raise ValueError("WGC GDT recycled-gold row was not found")
    for column, observation_date in recycling_columns.items():
        value = recycling_row[column] if column < len(recycling_row) else None
        if not finite(value):
            continue
        numeric = float(value)
        if numeric < 0:
            raise ValueError(f"Negative WGC GDT recycled-gold total for {observation_date}")
        recycling.append({"date": observation_date, "value": round(numeric, 10)})
    if not recycling:
        raise ValueError("WGC GDT workbook contained no quarterly recycled-gold supply")
    observations = {"demand": demand, "recycling": recycling}
    if cache_path is not None:
        parsed_cache = Path(cache_path)
        parsed_cache.parent.mkdir(parents=True, exist_ok=True)
        payload = {"source_file": source_path.name, "source_mtime_ns": source_stat.st_mtime_ns, "source_size": source_stat.st_size, "observations": observations}
        temporary = parsed_cache.with_suffix(parsed_cache.suffix + ".tmp")
        temporary.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
        temporary.replace(parsed_cache)
    return observations


def build_quarterly_output(
    variable_id: str,
    source_name: str,
    source_url: str,
    observations: list[dict[str, Any]],
    *,
    value_label: str,
    rising_signal: int,
    falling_signal: int,
    cached: bool = False,
    as_of_date: str | None = None,
) -> dict[str, Any]:
    """Build standard quarterly output with inapplicable short horizons."""
    ordered = sorted(observations, key=lambda item: str(item["date"]))
    current = ordered[-1] if ordered else None
    horizons: dict[str, Any] = {}
    for horizon in ("1-5d", "1-3m"):
        horizons[horizon] = {"signal": 0, "confidence": 1, "evidence": {"summary": f"Quarterly data does not support {horizon} horizon."}}
    for horizon, lookback in QUARTER_HORIZON_LOOKBACKS.items():
        if current is None or len(ordered) <= lookback:
            data = empty_data()
            if current is not None:
                data.update({"current_value": float(current["value"]), "current_date": str(current["date"])})
            summary = f"MISSING DATA — {lookback} prior quarterly observations are required; {max(0, len(ordered) - 1)} are available."
            if cached:
                summary += " SOURCE UNAVAILABLE — cached data used."
            horizons[horizon] = degraded(summary, data)
            continue
        comparison = ordered[-1 - lookback]
        current_value, comparison_value = float(current["value"]), float(comparison["value"])
        change = round(current_value - comparison_value, 10)
        change_pct = round(change / comparison_value * 100, 10) if comparison_value else None
        if change > 0:
            signal, direction = rising_signal, "rose"
        elif change < 0:
            signal, direction = falling_signal, "fell"
        else:
            signal, direction = 0, "was unchanged"
        pct_text = f"{change_pct:+.2f}%" if change_pct is not None else "percentage change unavailable"
        summary = f"{value_label} {direction} by {abs(change):.2f} tonnes ({pct_text}), {'bullish' if signal == 1 else 'bearish' if signal == -1 else 'neutral'} for gold."
        if cached:
            summary += " SOURCE UNAVAILABLE — cached data used."
        horizons[horizon] = {"signal": signal, "confidence": 1, "evidence": {"data": {"current_value": current_value, "current_date": str(current["date"]), "comparison_value": comparison_value, "comparison_date": str(comparison["date"]), "change_absolute": change, "change_pct": change_pct}, "summary": summary}}
    return {"variable_id": variable_id, "as_of_date": as_of_date or date.today().isoformat(), "source_name": source_name, "source_url": source_url, "data_frequency": "Quarterly", "observation_date": str(current["date"]) if current else None, "horizons": horizons}


def parse_official_holdings_workbook(path: Any) -> list[dict[str, Any]]:
    """Read the two canonical entity panels from the WGC PDF sheet."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    if "PDF" not in workbook.sheetnames:
        raise ValueError("WGC PDF sheet was not found")
    rows = list(workbook["PDF"].iter_rows(values_only=True))
    header_index = next((index for index, row in enumerate(rows) if {str(value).strip().lower() for value in row if value is not None} >= {"tonnes", "% of reserves**", "holdings as of"}), None)
    if header_index is None:
        raise ValueError("WGC official-holdings header was not found")
    records: list[dict[str, Any]] = []
    seen_countries: set[str] = set()
    for panel, name_index, tonnes_index, share_index, date_index, rank_index in (("left", 1, 2, 3, 4, 0), ("right", 6, 7, 8, 9, 5)):
        for row in rows[header_index + 1 :]:
            name = str(row[name_index]).strip() if name_index < len(row) and row[name_index] is not None else ""
            if not name or name.lower() in {"none", "nan", "total", "aggregate"}:
                break
            if name in seen_countries:
                continue
            if rank_index >= len(row) or not finite(row[rank_index]):
                break
            tonnes = row[tonnes_index] if tonnes_index < len(row) else None
            if not finite(tonnes) or float(tonnes) < 0:
                raise ValueError(f"Invalid official holdings for {name}")
            raw_share = row[share_index] if share_index < len(row) else None
            if isinstance(raw_share, str):
                raw_share = raw_share.split(")", 1)[-1].strip()
            if finite(raw_share) and not 0 <= float(raw_share) <= 1:
                raise ValueError(f"Invalid reserve share for {name}")
            share_value = float(raw_share) if finite(raw_share) else None
            source_date = row[date_index] if date_index < len(row) else None
            seen_countries.add(name)
            records.append({"country": name, "panel": panel, "date": parse_date(source_date), "source_date": str(source_date) if source_date is not None else None, "holdings": float(tonnes), "share": share_value})
    if not records:
        raise ValueError("WGC official-holdings sheet contained no valid records")
    return records


def empty_data() -> dict[str, Any]:
    """Return the common empty evidence data shape."""
    return {"current_value": None, "current_date": None, "comparison_value": None, "comparison_date": None, "change_absolute": None, "change_pct": None}


def degraded(summary: str, data: dict[str, Any] | None = None) -> dict[str, Any]:
    """Build a zero-confidence horizon result."""
    return {"signal": 0, "confidence": 0, "evidence": {"data": data or empty_data(), "summary": summary}}


def build_output(
    variable_id: str,
    source_name: str,
    source_url: str,
    observations: list[dict[str, Any]],
    *,
    cached: bool = False,
    as_of_date: str | None = None,
    value_label: str,
    rising_signal: int = 1,
    falling_signal: int = -1,
) -> dict[str, Any]:
    """Build standard four-horizon monthly output using rising/falling direction."""
    ordered = sorted(observations, key=lambda item: str(item["date"]))
    current = ordered[-1] if ordered else None
    horizons: dict[str, Any] = {}
    for horizon, lookback in HORIZON_LOOKBACKS.items():
        if current is None or len(ordered) < lookback:
            data = empty_data()
            if current is not None:
                data.update({"current_value": float(current["value"]), "current_date": str(current["date"])})
            summary = f"MISSING DATA — {lookback} valid monthly observations are required; {len(ordered)} are available."
            if cached:
                summary += " SOURCE UNAVAILABLE — cached data used."
            horizons[horizon] = degraded(summary, data)
            continue
        comparison = ordered[-lookback]
        current_value, comparison_value = float(current["value"]), float(comparison["value"])
        change = round(current_value - comparison_value, 10)
        change_pct = round(change / comparison_value * 100, 10) if comparison_value else None
        if change > 0:
            signal, direction = rising_signal, "rose"
        elif change < 0:
            signal, direction = falling_signal, "fell"
        else:
            signal, direction = 0, "was unchanged"
        percentage_text = f"{change_pct:+.2f}%" if change_pct is not None else "percentage change unavailable"
        summary = f"{value_label} {direction} by {abs(change):.2f} ({percentage_text}), {'bullish' if signal == 1 else 'bearish' if signal == -1 else 'neutral'} for gold."
        if cached:
            summary += " SOURCE UNAVAILABLE — cached data used."
        horizons[horizon] = {"signal": signal, "confidence": 1, "evidence": {"data": {"current_value": current_value, "current_date": str(current["date"]), "comparison_value": comparison_value, "comparison_date": str(comparison["date"]), "change_absolute": change, "change_pct": change_pct}, "summary": summary}}
    return {"variable_id": variable_id, "as_of_date": as_of_date or date.today().isoformat(), "source_name": source_name, "source_url": source_url, "data_frequency": "Monthly", "observation_date": str(current["date"]) if current else None, "horizons": horizons}
