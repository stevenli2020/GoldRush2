"""Extract the WGC Chinese premium/discount series (L9-001)."""

from __future__ import annotations

import json
import math
from datetime import date, datetime
from pathlib import Path
from typing import Any

from goldrush2.collectors.wgc import LAST_FETCH_STALE, LAST_FETCH_USED_CACHE, fetch_wgc_gold_premiums
from goldrush2.extractors._wgc_common import parse_date

VARIABLE_ID = "L9-001"
SOURCE_NAME = "WGC - China Physical Gold Premium/Discount (USD/oz)"
SOURCE_URL = "https://www.gold.org/goldhub/data/gold-premium"
RAW_PATH = Path("data/raw/wgc/gold-premiums.xlsx")
PARSED_CACHE_PATH = Path("data/cache/wgc/gold_premiums.json")
OUTPUT_PATH = Path("data/current/L9-001.json")
SHEET_NAME = "Chinese premiums-discounts"
LOOKBACKS = {"1-5d": 5, "1-3m": 63, "1-3y": 252, "3-10y": 756}


def _number(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _date(value: Any) -> str | None:
    return parse_date(value)


def parse_premiums_workbook(path: Path, *, cache_path: Path | None = PARSED_CACHE_PATH) -> list[dict[str, Any]]:
    """Parse and validate the exact WGC China sheet.

    WGC currently publishes only the USD/oz spread.  If a future workbook
    adds local and international price columns, ``premium_pct`` is calculated
    from those columns; otherwise it remains null and USD/oz is the source
    signal basis rather than fabricating an international price.
    """
    path = Path(path)
    stat = path.stat()
    if cache_path is not None:
        try:
            cached = json.loads(Path(cache_path).read_text(encoding="utf-8"))
            if cached.get("source_file") == path.name and cached.get("source_mtime_ns") == stat.st_mtime_ns and cached.get("source_size") == stat.st_size and isinstance(cached.get("observations"), list):
                return cached["observations"]
        except (OSError, ValueError, json.JSONDecodeError):
            pass

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"required sheet missing: {SHEET_NAME}")
    rows = list(workbook[SHEET_NAME].iter_rows(values_only=True))
    title = " ".join(str(v).strip() for row in rows[:5] for v in row if v is not None)
    title_lower = title.lower()
    if "chinese premium/discount" not in title_lower or "us$/oz" not in title_lower or "5 day moving average" not in title_lower:
        raise ValueError("Chinese sheet title does not verify series, unit, and moving-average definition")

    observations: dict[str, dict[str, Any]] = {}
    for row in rows[5:]:
        if not row or all(value is None for value in row):
            continue
        observed = _date(row[0] if len(row) else None)
        if observed is None:
            raise ValueError(f"invalid observation date: {row[0] if row else None}")
        premium_usd = _number(row[1] if len(row) > 1 else None)
        if premium_usd is None:
            raise ValueError(f"invalid Chinese premium/discount value for {observed}")
        local = _number(row[2]) if len(row) > 2 else None
        international = _number(row[3]) if len(row) > 3 else None
        premium_pct = None
        if local is not None and international is not None:
            if international == 0:
                raise ValueError(f"international price is zero for {observed}")
            premium_pct = (local - international) / international * 100
            if not math.isfinite(premium_pct):
                raise ValueError(f"non-finite premium percentage for {observed}")
        record = {"date": observed, "value": premium_usd, "premium_usd": premium_usd, "premium_pct": premium_pct}
        if observed in observations:
            raise ValueError(f"duplicate observation date: {observed}")
        observations[observed] = record
    result = [observations[key] for key in sorted(observations)]
    if not result:
        raise ValueError("workbook contains no Chinese premium/discount observations")
    if cache_path is not None:
        target = Path(cache_path)
        target.parent.mkdir(parents=True, exist_ok=True)
        payload = {"source_file": path.name, "source_mtime_ns": stat.st_mtime_ns, "source_size": stat.st_size, "observations": result}
        temporary = target.with_suffix(target.suffix + ".tmp")
        temporary.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
        temporary.replace(target)
    return result


def _level(value: float) -> str:
    if value > 0:
        return "positive"
    if value < 0:
        return "negative"
    return "near_zero"


def build_output(observations: list[dict[str, Any]], *, cached: bool = False, as_of_date: str | None = None) -> dict[str, Any]:
    ordered = sorted(observations, key=lambda item: str(item["date"]))
    current = ordered[-1] if ordered else None
    horizons: dict[str, Any] = {}
    for horizon, lookback in LOOKBACKS.items():
        if current is None or len(ordered) <= lookback:
            data = {"current_value_pct": current.get("premium_pct") if current else None, "current_value_usd": current.get("premium_usd") if current else None, "current_date": current.get("date") if current else None, "comparison_value_pct": None, "comparison_date": None, "change_pct": None, "absolute_level": _level(float(current["premium_usd"])) if current else None}
            summary = f"MISSING DATA — {lookback} prior weekly observations are required; {max(0, len(ordered) - 1)} are available."
            if cached or LAST_FETCH_STALE:
                summary += " SOURCE UNAVAILABLE — cached data used."
            horizons[horizon] = {"signal": 0, "confidence": 0, "evidence": {"data": data, "summary": summary}}
            continue
        comparison = ordered[-1 - lookback]
        current_value = current.get("premium_pct")
        comparison_value = comparison.get("premium_pct")
        basis = "percentage" if current_value is not None and comparison_value is not None else "USD/oz"
        if basis == "USD/oz":
            current_value, comparison_value = float(current["premium_usd"]), float(comparison["premium_usd"])
        change = round(float(current_value) - float(comparison_value), 10)
        signal = 1 if change > 0 else -1 if change < 0 else 0
        direction = "widened" if signal == 1 else "narrowed" if signal == -1 else "was unchanged"
        data = {"current_value_pct": current.get("premium_pct"), "current_value_usd": float(current["premium_usd"]), "current_date": current["date"], "comparison_value_pct": comparison.get("premium_pct"), "comparison_date": comparison["date"], "change_pct": change, "absolute_level": _level(float(current["premium_usd"]))}
        summary = f"China gold premium {direction} by {abs(change):.2f} {('percentage points' if basis == 'percentage' else 'USD/oz')} compared to {lookback} weeks ago, indicating {'strengthening physical demand, bullish' if signal == 1 else 'weakening physical demand, bearish' if signal == -1 else 'stable physical demand, neutral'} for gold. Current premium: {float(current['premium_usd']):+.2f} USD/oz."
        if basis == "USD/oz":
            summary += " Percentage unavailable in the WGC workbook; USD/oz used as the signal basis."
        if cached or LAST_FETCH_USED_CACHE:
            summary += " SOURCE UNAVAILABLE — cached data used."
        horizons[horizon] = {"signal": signal, "confidence": 1, "evidence": {"data": data, "summary": summary}}
    return {"variable_id": VARIABLE_ID, "data_frequency": "Weekly", "source_name": SOURCE_NAME, "source_url": "https://www.gold.org/", "observation_date": current["date"] if current else None, "as_of_date": as_of_date or date.today().isoformat(), "horizons": horizons}


def run(*, workbook: Path | None = None, output_path: Path = OUTPUT_PATH, raw_dir: Path = Path("data/raw/wgc")) -> dict[str, Any]:
    source = Path(workbook) if workbook else fetch_wgc_gold_premiums(raw_dir)
    cached = LAST_FETCH_USED_CACHE
    if source is None:
        # Preserve a stale raw workbook for an explicit degraded output when
        # the authenticated source is unavailable.
        fallback = Path(raw_dir) / "gold-premiums.xlsx"
        if fallback.exists():
            source, cached = fallback, True
    if source is None:
        raise RuntimeError("WGC gold-premiums workbook is unavailable")
    result = build_output(parse_premiums_workbook(source), cached=cached)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = output_path.with_suffix(output_path.suffix + ".tmp")
    temporary.write_text(json.dumps(result, indent=2) + "\n", encoding="utf-8")
    temporary.replace(output_path)
    return result


if __name__ == "__main__":
    run()
